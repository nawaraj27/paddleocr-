"""Export extracted documents to CSV / Excel / JSON."""
from __future__ import annotations

import csv
import io
import json
from typing import Any

CONTENT_TYPES = {
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "json": "application/json",
}

_COLUMNS = ["invoice_number", "vendor_name", "invoice_date", "currency",
            "subtotal", "tax", "total_amount", "category", "confidence",
            "is_saved", "created_at"]


def _row(doc) -> dict[str, Any]:
    return {
        "invoice_number": doc.invoice_number,
        "vendor_name": doc.vendor_name,
        "invoice_date": doc.invoice_date.isoformat() if doc.invoice_date else "",
        "currency": doc.currency,
        "subtotal": doc.subtotal,
        "tax": doc.tax,
        "total_amount": doc.total_amount,
        "category": doc.category.name if doc.category_id else "",
        "confidence": round(doc.confidence, 3),
        "is_saved": doc.is_saved,
        "created_at": doc.created_at.isoformat() if doc.created_at else "",
    }


def export_documents(docs, fmt: str) -> tuple[bytes, str]:
    if fmt == "csv":
        return _csv(docs), "documents.csv"
    if fmt == "json":
        return _json(docs), "documents.json"
    if fmt == "xlsx":
        return _xlsx(docs), "documents.xlsx"
    raise ValueError(f"unknown format {fmt}")


def _csv(docs) -> bytes:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=_COLUMNS)
    w.writeheader()
    for d in docs:
        w.writerow(_row(d))
    return buf.getvalue().encode("utf-8")


def _json(docs) -> bytes:
    payload = []
    for d in docs:
        row = _row(d)
        row["items"] = [
            {"name": i.name, "quantity": float(i.quantity or 0),
             "unit_price": float(i.unit_price) if i.unit_price is not None else None,
             "amount": float(i.amount) if i.amount is not None else None}
            for i in d.items.all()
        ]
        payload.append(row)
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


def _xlsx(docs) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"
    ws.append([c.replace("_", " ").title() for c in _COLUMNS])
    head_fill = PatternFill("solid", fgColor="D4E4B4")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = head_fill
    for d in docs:
        r = _row(d)
        ws.append([r[c] for c in _COLUMNS])
    for i, c in enumerate(_COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18

    ws2 = wb.create_sheet("Items")
    ws2.append(["Invoice", "Vendor", "Item", "Qty", "Unit Price", "Amount"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = head_fill
    for d in docs:
        for it in d.items.all():
            ws2.append([d.invoice_number, d.vendor_name, it.name,
                        float(it.quantity or 0),
                        float(it.unit_price) if it.unit_price is not None else None,
                        float(it.amount) if it.amount is not None else None])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
